from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Brand Contacts"

headers = ["Brand Name", "Category", "Products", "Website", "Instagram", "Email", "Why Partner", "Notes"]
header_fill = PatternFill("solid", fgColor="00FF00")
header_font = Font(name="Arial", bold=True, size=11, color="000000")
thin_border = Border(
    left=Side(style="thin", color="333333"),
    right=Side(style="thin", color="333333"),
    top=Side(style="thin", color="333333"),
    bottom=Side(style="thin", color="333333"),
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 24
ws.column_dimensions["D"].width = 30
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 32
ws.column_dimensions["G"].width = 50
ws.column_dimensions["H"].width = 20

brands = [
    ("Mood Hoops", "Hoop Companies", "LED Smart Hoops", "https://moodhoops.com", "@moodhoops", "info@moodhoops.com", "Industry leader. Large ambassador program."),
    ("Astral Hoops", "Hoop Companies", "LED & Poly Hoops", "https://astralhoops.com", "@astralhoops", "support@astralhoops.com", "Premium smart hoops with app control."),
    ("Hyperion Hoops", "Hoop Companies", "LED Hoops", "https://hyperionhoops.com", "@hyperionhoops", "info@hyperionhoops.com", "High-end LED hoops. Growing brand."),
    ("The Spinsterz", "Hoop Companies", "Hoops & Flow Props", "https://thespinsterz.com", "@thespinsterz", "info@thespinsterz.com", "Massive catalog. Active sponsorship program."),
    ("Ruby Hooping", "Hoop Companies", "Handmade Hoops", "https://rubyhooping.com", "@rubyhooping", "hello@rubyhooping.com", "Artisan hoop maker."),
    ("Hoop Mamas", "Hoop Companies", "Custom Hoops", "https://hoopmamas.com", "@hoopmamas", "info@hoopmamas.com", "Community-focused brand."),
    ("Cosmic Hoops", "Hoop Companies", "LED & Polypro Hoops", "https://cosmichoops.com", "@cosmichoops", "", "Budget-friendly LEDs."),
    ("Sacred Flow Art", "Hoop Companies", "Hoops & Props", "https://sacredflowart.com", "@sacredflowart", "info@sacredflowart.com", "Multi-prop brand."),
    ("Flowtoys", "Poi & Staff", "LED Poi, Staffs, Clubs", "https://flowtoys.com", "@flowtoys", "info@flowtoys.com", "Gold standard in flow props."),
    ("Home of Poi", "Poi & Staff", "Poi & Fire Props", "https://homeofpoi.com", "@homeofpoi", "info@homeofpoi.com", "New Zealand-based, global reach."),
    ("Dark Monk", "Poi & Staff", "Fire Poi & Props", "https://darkmonk.com", "@darkmonkdesign", "info@darkmonk.com", "Premium fire props."),
    ("Pyroterra", "Poi & Staff", "Fire & LED Props", "https://pyroterra.cz", "@pyroterra", "info@pyroterra.cz", "European brand expanding globally."),
    ("Trick Concepts", "Poi & Staff", "Contact Poi & Props", "https://trickconcepts.com", "@trickconcepts", "info@trickconcepts.com", "Innovation-focused brand."),
    ("Spinballs", "Poi & Staff", "LED Poi & Accessories", "https://spinballs.com", "@spinballs", "hello@spinballs.com", "Affordable LED poi. Sponsors many artists."),
    ("Flow on Fire", "Poi & Staff", "Fire Props", "https://flowonfire.co.uk", "@flowonfire", "", "Artisan fire prop maker."),
    ("UltraPoi", "Poi & Staff", "LED Poi", "https://ultrapoi.com", "@ultrapoi", "support@ultrapoi.com", "Feature-rich LED poi with app."),
    ("Emazing Lights", "LED & Gloving", "LED Gloves & Orbits", "https://emazinglights.com", "@emazinglights", "team@emazinglights.com", "Dominant in gloving scene."),
    ("Future Crew", "LED & Gloving", "LED Gloves", "https://futurecrew.com", "@futurecrewlights", "info@futurecrew.com", "Premium gloving brand."),
    ("GloFX", "LED & Gloving", "LED Glasses & Accessories", "https://glofx.com", "@glofx", "info@glofx.com", "EDM accessories brand."),
    ("Lux LED", "LED & Gloving", "LED Flow Props", "https://luxledlights.com", "@luxledlights", "", "Boutique LED manufacturer."),
    ("Neon Nightlife", "LED & Gloving", "LED Party Supplies", "https://neonnightlife.com", "@neonnightlife", "support@neonnightlife.com", "Broad LED product line."),
    ("Electro Glow", "LED & Gloving", "LED Wearables", "https://electroglow.com", "@electroglow", "", "Growing LED fashion brand."),
    ("Flow DNA", "Flow Toys & Props", "Contact Staffs & Wands", "https://flowdna.co.uk", "@flowdna", "", "Premium contact props."),
    ("Flames N Games", "Flow Toys & Props", "Juggling & Flow Props", "https://flamesnagames.co.uk", "@flamesnagames", "info@flamesnagames.co.uk", "UK-based, global distribution."),
    ("Wizard Props", "Flow Toys & Props", "Levitation Wands", "", "@wizardprops", "", "Specialized levi wand maker."),
    ("Crystal Levity", "Flow Toys & Props", "Crystal Balls & Contact", "https://crystallevity.com", "@crystallevity", "", "Contact juggling specialists."),
    ("Moodhoops (Fans)", "Flow Toys & Props", "LED Fans", "https://moodhoops.com/fans", "@moodhoops", "info@moodhoops.com", "Expanding into fan dance."),
    ("Ignis Pixel", "Flow Toys & Props", "LED Visual Poi", "https://ignispixel.com", "@ignispixel", "info@ignispixel.com", "Pixel poi with custom images."),
    ("DrexFactor", "Flow Toys & Props", "Instructional Content", "https://drexfactor.com", "@drexfactor", "drex@drexfactor.com", "Key influencer. Collab builds credibility."),
    ("Play Juggling", "Flow Toys & Props", "European Flow Props", "https://playjuggling.com", "@playjuggling", "info@playjuggling.com", "Italian brand. Growing US presence."),
    ("iHeartRaves", "Festival Fashion", "Festival Fashion", "https://iheartraves.com", "@iheartraves", "ambassadors@iheartraves.com", "Major festival clothing brand."),
    ("INTO THE AM", "Festival Fashion", "Festival Apparel", "https://intotheam.com", "@intotheam", "ambassadors@intotheam.com", "EDM-focused clothing."),
    ("Rave Wonderland", "Festival Fashion", "Rave Accessories", "https://ravewonderland.com", "@ravewonderland", "info@ravewonderland.com", "Accessories and costumes."),
    ("Freedom Rave Wear", "Festival Fashion", "Festival Outfits", "https://freedomravewear.com", "@freedomravewear", "info@freedomravewear.com", "Body-positive rave fashion."),
    ("Lunafide", "Festival Fashion", "Festival Clothing", "https://lunafide.com", "@lunafide", "hello@lunafide.com", "Premium festival fashion."),
    ("Electric Family", "Festival Fashion", "Bracelets & Apparel", "https://electricfamily.com", "@electricfamily", "info@electricfamily.com", "Community-driven brand."),
    ("SoJourner Bags", "Festival Fashion", "Hydration Packs", "https://sojournersbags.com", "@sojournerbags", "hello@sojournersbags.com", "Practical festival gear."),
    ("Lunchbox Packs", "Festival Fashion", "Hydration Packs", "https://lunchboxpacks.com", "@lunchboxpacks", "", "LED hydration packs."),
    ("Manduka", "Wellness & Lifestyle", "Yoga Mats & Gear", "https://manduka.com", "@manduka", "ambassadors@manduka.com", "Premium yoga brand."),
    ("Vuori", "Wellness & Lifestyle", "Performance Apparel", "https://vuoriclothing.com", "@vuoriclothing", "ambassadors@vuoriclothing.com", "Athleisure brand."),
    ("Liquid IV", "Wellness & Lifestyle", "Hydration Products", "https://liquid-iv.com", "@liquidiv", "partnerships@liquid-iv.com", "Festival staple."),
    ("CLIF Bar", "Wellness & Lifestyle", "Energy Bars", "https://clifbar.com", "@clifbar", "ambassadors@clifbar.com", "Active lifestyle brand."),
    ("Hydro Flask", "Wellness & Lifestyle", "Water Bottles", "https://hydroflask.com", "@hydroflask", "partnerships@hydroflask.com", "Festival essential."),
    ("Blenders Eyewear", "Wellness & Lifestyle", "Sunglasses", "https://blenderseyewear.com", "@blenderseyewear", "ambassadors@blenderseyewear.com", "Festival-friendly sunglasses."),
]

data_font = Font(name="Arial", size=10)
cat_colors = {
    "Hoop Companies": PatternFill("solid", fgColor="E8FFE8"),
    "Poi & Staff": PatternFill("solid", fgColor="FFE8FF"),
    "LED & Gloving": PatternFill("solid", fgColor="E8FFE8"),
    "Flow Toys & Props": PatternFill("solid", fgColor="FFE8FF"),
    "Festival Fashion": PatternFill("solid", fgColor="E8FFE8"),
    "Wellness & Lifestyle": PatternFill("solid", fgColor="FFE8FF"),
}

for i, b in enumerate(brands, 2):
    for col, val in enumerate(b, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font = data_font
        cell.border = thin_border
        if b[1] in cat_colors:
            cell.fill = cat_colors[b[1]]
    ws.cell(row=i, column=8, value="").font = data_font
    ws.cell(row=i, column=8).border = thin_border

ws.auto_filter.ref = f"A1:H{len(brands)+1}"
ws.freeze_panes = "A2"

wb.save("C:/cash/public/brand-contacts.xlsx")
print(f"Created brand-contacts.xlsx with {len(brands)} brands")
